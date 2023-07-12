from io import BytesIO
from tempfile import NamedTemporaryFile
from typing import NamedTuple

from django.contrib.auth import get_user_model
from django.utils import timezone
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

from rooms.async_db import (
    get_final_user_beer_ratings,
    get_final_beers_ratings
)

User = get_user_model()

MAIN_HEADERS = [
    'NUMER', 'KOLOR', 'PIANA', 'ZAPACH', 'SMAK', 'OPINIA', 'OCENA KOŃCOWA'
]

SECOND_HEADERS = [
    'NUMER', 'NAZWA', 'BROWAR', 'STYL', 'ŚREDNIA OCENA'
]


class UserRatingRow(NamedTuple):
    number: int
    color: int
    foam: int
    smell: int
    taste: int
    opinion: str
    note: float


class BeerRatingRow(NamedTuple):
    number: int
    name: str
    brewery: str
    style: str
    average_rating: float


def get_formatted_user_rating_row(number: int, rating: dict) -> UserRatingRow:
    return UserRatingRow(
        number=number,
        color=rating['color'],
        foam=rating['foam'],
        smell=rating['smell'],
        taste=rating['taste'],
        opinion=rating['opinion'],
        note=rating['note']
    )


def get_formatted_beer_rating_row(number: int, rating: dict) -> BeerRatingRow:
    beer = rating['beer']
    return BeerRatingRow(
        number=number,
        name=beer['name'],
        brewery=beer['brewery'],
        style=beer['style'],
        average_rating=rating['average_rating']
    )


def format_columns_width(sheet, headers: list[str]) -> None:
    for index, _ in enumerate(headers):
        col_letter = get_column_letter(index + 1)
        col_width = 20

        # index column
        if index == 0:
            col_width = 10

        sheet.column_dimensions[col_letter].width = col_width


def collect_user_ratings(room_name: str, user: User) -> list[UserRatingRow]:
    user_ratings = get_final_user_beer_ratings(room_name=room_name, user=user)

    ratings = []
    for index, rating in enumerate(user_ratings):
        row = get_formatted_user_rating_row(index + 1, rating)
        ratings.append(row)
    return ratings


def collect_beer_ratings(room_name) -> list[BeerRatingRow]:
    final_ratings = get_final_beers_ratings(room_name)

    ratings = []
    for index, rating in enumerate(final_ratings):
        row = get_formatted_beer_rating_row(index + 1, rating)
        ratings.append(row)
    return ratings


def generate_excel_report(room_name: str, user: User) -> BytesIO:
    now = timezone.now()
    wb = Workbook()

    # main sheet with user's ratings
    main_sheet = wb.active
    main_sheet.title = 'Twoje oceny'
    main_sheet.append(MAIN_HEADERS)
    format_columns_width(main_sheet, MAIN_HEADERS)
    user_ratings = collect_user_ratings(room_name=room_name, user=user)
    for user_rating in user_ratings:
        main_sheet.append(user_rating)

    # second sheet with results (average ratings)
    second_sheet = wb.create_sheet('Podsumowanie')
    second_sheet.append(SECOND_HEADERS)
    format_columns_width(second_sheet, SECOND_HEADERS)
    final_ratings = collect_beer_ratings(room_name=room_name)
    for final_rating in final_ratings:
        second_sheet.append(final_rating)

    # save to temporary file and return it as a buffer
    with NamedTemporaryFile() as tmp:
        wb.save(tmp.name)
        output = BytesIO(tmp.read())
        return output
